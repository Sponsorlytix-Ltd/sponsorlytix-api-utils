from openpyxl import load_workbook
import json



country = ''




class BroadcastSheets:
    def __init__(self):
        self.country = ''
    
    def __builder_broadcast(self, row_values):
        if row_values[0]:
            self.country = row_values[0]
        return {
            'territory': '',
            'country': self.country,
            'broadcast_name': row_values[1],
            'price': row_values[2],
            'average_last_month': row_values[3],
            'website': row_values[4],
            'ott_subs': row_values[5],
            'otts_subs_prices': row_values[6]
        }

    def __map_broadcast(self, row):
        row_values = [column_row.value for column_row in row]
        return self.__builder_broadcast(row_values)


    def process_broadcast_sheet(self, sheet):
        sheet_data = map(self.__map_broadcast, sheet.iter_rows(min_row=2))
        return sheet_data

    def process_broadcast(self):
        workbook = load_workbook(filename="sheets/Broadcasters Rate Cards.xlsx")
        broadcast_data = dict()
        for territory in workbook.sheetnames:
            sheet = workbook[territory]
            broadcast_data.update(
                {
                    territory: self.process_broadcast_sheet(sheet)
                }
            )

BroadcastSheets().process_broadcast()
